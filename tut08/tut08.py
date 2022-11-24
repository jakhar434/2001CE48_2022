#####importing libraries

import openpyxl

import pandas as pd
import os
from datetime import datetime

start_time = datetime.now()

ind_inning = open("india_inns2.txt", "r+")
pak_inning = open("pak_inns1.txt", "r+")

teams = open("teams.txt", "r+")

input_team = teams.readlines()
##########################################################
# paistan team
team_of_pakistan = input_team[0]
pak_players_list = team_of_pakistan[23:-1:].split(",")
# indian team
team_of_india = input_team[2]
ind_players_list = team_of_india[20:-1:].split(",")


lst_ind = ind_inning.readlines()
for i in lst_ind:
    if i == '\n':
        lst_ind.remove(i)


lst_pak = pak_inning.readlines()  # 123
for i in lst_pak:
    if i == '\n':
        lst_pak.remove(i)

wb = openpyxl.Workbook()
sheet = wb.active



out_pak_bat = {}
ind_bowlers = {}

indian_second_bat = {}
pak_bats = {}

pak_bowlers = {}

pak_byes = 0
pak_bowlers_total = 0


indian_fall_of_wickets = 0  # number of wicket of india team
pakistan_fall_of_wickets = 0  # number of wicket of pakistan team


for l in lst_pak:
    x = l.index(".")
    over_pak = l[0:x+2]
    flag = l[x+2::].split(",")
    curr_ball = flag[0].split("to")  # 0 2

    if f"{curr_ball[0].strip()}" not in ind_bowlers.keys():
        ind_bowlers[f"{curr_ball[0].strip()}"] = [1, 0, 0, 0, 0, 0, 0]

    elif "wide" in flag[1]:
        pass

    elif "bye" in flag[1]:
        if "FOUR" in flag[2]:
            pak_byes += 4
        elif "1" in flag[2]:
            pak_byes += 1
        elif "2" in flag[2]:
            pak_byes += 2
        elif "3" in flag[2]:
            pak_byes += 3
        elif "4" in flag[2]:
            pak_byes += 4
        elif "5" in flag[2]:
            pak_byes += 5

    else:
        ind_bowlers[f"{curr_ball[0].strip()}"][0] += 1

    if f"{curr_ball[1].strip()}" not in pak_bats.keys() and flag[1] != "wide":
        pak_bats[f"{curr_ball[1].strip()}"] = [0, 1, 0, 0, 0]
    elif "wide" in flag[1]:
        pass
    else:
        pak_bats[f"{curr_ball[1].strip()}"][1] += 1

    if "out" in flag[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][3] += 1
        if "Bowled" in flag[1].split("!!")[0]:
            out_pak_bat[f"{curr_ball[1].strip()}"] = ("b" + curr_ball[0])
        elif "Caught" in flag[1].split("!!")[0]:
            w = (flag[1].split("!!")[0]).split("by")
            out_pak_bat[f"{curr_ball[1].strip()}"] = (
                "c" + w[1] + " b " + curr_ball[0])
        elif "Lbw" in flag[1].split("!!")[0]:
            out_pak_bat[f"{curr_ball[1].strip()}"] = ("lbw  b "+curr_ball[0])

    if "no run" in flag[1] or "out" in flag[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 0
        pak_bats[f"{curr_ball[1].strip()}"][0] += 0

    elif "1 run" in flag[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 1
        pak_bats[f"{curr_ball[1].strip()}"][0] += 1

    elif "2 run" in flag[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 2
        pak_bats[f"{curr_ball[1].strip()}"][0] += 2
    elif "3 run" in flag[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 3
        pak_bats[f"{curr_ball[1].strip()}"][0] += 3

    elif "4 run" in flag[1]:  ##########there is no boundary but player get 4 run###########
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        pak_bats[f"{curr_ball[1].strip()}"][0] += 4

    elif "FOUR" in flag[1]:
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        pak_bats[f"{curr_ball[1].strip()}"][0] += 4
        pak_bats[f"{curr_ball[1].strip()}"][2] += 1

    elif "SIX" in flag[1]:  ######### for six###############
        ind_bowlers[f"{curr_ball[0].strip()}"][2] += 6
        pak_bats[f"{curr_ball[1].strip()}"][0] += 6
        pak_bats[f"{curr_ball[1].strip()}"][3] += 1

    elif "wide" in flag[1]: ######### for wide balls##########
        if "wides" in flag[1]:

            ind_bowlers[f"{curr_ball[0].strip()}"][2] += int(flag[1][1])
            ind_bowlers[f"{curr_ball[0].strip()}"][5] += int(flag[1][1])

        else:
            ind_bowlers[f"{curr_ball[0].strip()}"][2] += 1
            ind_bowlers[f"{curr_ball[0].strip()}"][5] += 1

for values in pak_bats.values():
    values[-1] = round((values[0]/values[1])*100, 2)


ind_bowlers_total = 0
ind_byes = 0
out_ind_bat = {}

####################################################################
# INDIA INNING
for l in lst_ind:
    x = l.index(".")
    over_ind = l[0:x+2]

    flag = l[x+2::].split(",")

    curr_ball = flag[0].split("to")
    if f"{curr_ball[0].strip()}" not in pak_bowlers.keys():
        pak_bowlers[f"{curr_ball[0].strip()}"] = [1, 0, 0, 0, 0, 0, 0]
    elif "wide" in flag[1]:
        pass
    elif "bye" in flag[1]:
        if "FOUR" in flag[2]:
            ind_byes += 4
        elif "1" in flag[2]:
            ind_byes += 1
        elif "2" in flag[2]:
            ind_byes += 2
        elif "3" in flag[2]:
            ind_byes += 3
        elif "4" in flag[2]:
            ind_byes += 4
        elif "5" in flag[2]:
            ind_byes += 5
    else:
        pak_bowlers[f"{curr_ball[0].strip()}"][0] += 1

    if f"{curr_ball[1].strip()}" not in indian_second_bat.keys() and flag[1] != "wide":
        indian_second_bat[f"{curr_ball[1].strip()}"] = [
            0, 1, 0, 0, 0]  # [runs,ball,4s,6s,sr]
    elif "wide" in flag[1]:
        pass
    else:
        indian_second_bat[f"{curr_ball[1].strip()}"][1] += 1

    if "out" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][3] += 1

        if "Bowled" in flag[1].split("!!")[0]:
            out_ind_bat[f"{curr_ball[1].strip()}"] = ("b" + curr_ball[0])
        elif "Caught" in flag[1].split("!!")[0]:
            w = (flag[1].split("!!")[0]).split("by")
            out_ind_bat[f"{curr_ball[1].strip()}"] = (
                "c" + w[1] + " b " + curr_ball[0])
        elif "Lbw" in flag[1].split("!!")[0]:
            out_ind_bat[f"{curr_ball[1].strip()}"] = ("lbw  b "+curr_ball[0])




    if "no run" in flag[1] or "out" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 0
        indian_second_bat[f"{curr_ball[1].strip()}"][0] += 0
    elif "1 run" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 1
        indian_second_bat[f"{curr_ball[1].strip()}"][0] += 1
    elif "2 run" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 2
        indian_second_bat[f"{curr_ball[1].strip()}"][0] += 2
    elif "3 run" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 3
        indian_second_bat[f"{curr_ball[1].strip()}"][0] += 3
    elif "4 run" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        indian_second_bat[f"{curr_ball[1].strip()}"][0] += 4
    elif "FOUR" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 4
        indian_second_bat[f"{curr_ball[1].strip()}"][0] += 4
        indian_second_bat[f"{curr_ball[1].strip()}"][2] += 1
    elif "SIX" in flag[1]:
        pak_bowlers[f"{curr_ball[0].strip()}"][2] += 6
        indian_second_bat[f"{curr_ball[1].strip()}"][0] += 6
        indian_second_bat[f"{curr_ball[1].strip()}"][3] += 1
    elif "wide" in flag[1]:
        if "wides" in flag[1]:
            pak_bowlers[f"{curr_ball[0].strip()}"][2] += int(flag[1][1])
            pak_bowlers[f"{curr_ball[0].strip()}"][5] += int(flag[1][1])
        else:
            pak_bowlers[f"{curr_ball[0].strip()}"][2] += 1
            pak_bowlers[f"{curr_ball[0].strip()}"][5] += 1


for values in indian_second_bat.values():
    values[-1] = round((values[0]/values[1])*100, 2)

for values in pak_bats.values():
    values[-1] = round((values[0]/values[1])*100, 2)

for values in ind_bowlers.values():
    if values[0] % 6 == 0:
        values[0] = values[0]//6
    else:
        values[0] = (values[0]//6) + (values[0] % 6)/10

for values in pak_bowlers.values():
    if values[0] % 6 == 0:
        values[0] = values[0]//6
    else:
        values[0] = (values[0]//6) + (values[0] % 6)/10

for values in ind_bowlers.values():
    x = str(values[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        values[-1] = round((values[2]/balls)*6, 1)
    else:
        values[-1] = round((values[2]/values[0]), 1)


for values in pak_bowlers.values():
    x = str(values[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        values[-1] = round((values[2]/balls)*6, 1)
    else:
        values[-1] = round((values[2]/values[0]), 1)


###########################################################################
# PAKISTAN BATTING
pakistan_batsman_names = []
for count in pak_bats.keys():
    pakistan_batsman_names.append(count)


for i in range(len(pak_bats)):
    sheet.cell(5+i, 1).value = pakistan_batsman_names[i]
    sheet.cell(5+i, 5).value = pak_bats[pakistan_batsman_names[i]][0]
    sheet.cell(5+i, 6).value = pak_bats[pakistan_batsman_names[i]][1]
    sheet.cell(5+i, 7).value = pak_bats[pakistan_batsman_names[i]][2]
    sheet.cell(5+i, 8).value = pak_bats[pakistan_batsman_names[i]][3]
    sheet.cell(5+i, 9).value = pak_bats[pakistan_batsman_names[i]][4]

    if pakistan_batsman_names[i] not in out_pak_bat:
        sheet.cell(5+i, 3).value = "not out"
    else:
        sheet.cell(5+i, 3).value = out_pak_bat[pakistan_batsman_names[i]]

sheet.cell(3, 1).value = "BATTERS"
sheet["E3"] = "RUNS"
sheet["F3"] = "BALLS"
sheet["G3"] = " 4s "
sheet["H3"] = " 6s "
sheet["I3"] = "  SR  "

#########################################################
# INDIA BOWLING

sheet["A18"] = "BOWLER"
sheet["C18"] = "OVER"
sheet["D18"] = "MAIDEN"
sheet["E18"] = "RUNS"
sheet["F18"] = "WICKET"
sheet["G18"] = "NO-BALL"
sheet["H18"] = "WIDE"
sheet["I18"] = "ECONOMY"

names_pak_bowlers = []
for count in pak_bowlers.keys():
    names_pak_bowlers.append(count)

for i in range(len(pak_bowlers)):
    sheet.cell(42+i, 1).value = names_pak_bowlers[i]
    sheet.cell(42+i, 3).value = pak_bowlers[names_pak_bowlers[i]][0]
    sheet.cell(42+i, 4).value = pak_bowlers[names_pak_bowlers[i]][1]
    sheet.cell(42+i, 5).value = pak_bowlers[names_pak_bowlers[i]][2]
    sheet.cell(42+i, 6).value = pak_bowlers[names_pak_bowlers[i]][3]
    sheet.cell(42+i, 7).value = pak_bowlers[names_pak_bowlers[i]][4]
    sheet.cell(42+i, 8).value = pak_bowlers[names_pak_bowlers[i]][5]
    sheet.cell(42+i, 9).value = pak_bowlers[names_pak_bowlers[i]][6]
    pak_bowlers_total += pak_bowlers[names_pak_bowlers[i]][2]
    indian_fall_of_wickets += pak_bowlers[names_pak_bowlers[i]][3]

###############################################
# INDIA BATTING
sheet.cell(11+len(pak_bats)+len(pak_bowlers), 1).value = "# INDIA"
sheet.cell(11+len(pak_bats)+len(pak_bowlers), 2).value = " INNINGS"

names_ind_batters = []
for count in indian_second_bat.keys():
    names_ind_batters.append(count)


for i in range(len(indian_second_bat)):
    sheet.cell(31+i, 1).value = names_ind_batters[i]
    sheet.cell(31+i, 5).value = indian_second_bat[names_ind_batters[i]][0]
    sheet.cell(31+i, 6).value = indian_second_bat[names_ind_batters[i]][1]
    sheet.cell(31+i, 7).value = indian_second_bat[names_ind_batters[i]][2]
    sheet.cell(31+i, 8).value = indian_second_bat[names_ind_batters[i]][3]
    sheet.cell(31+i, 9).value = indian_second_bat[names_ind_batters[i]][4]

    if names_ind_batters[i] not in out_ind_bat:
        sheet.cell(31+i, 3).value = "not out"
    else:
        sheet.cell(31+i, 3).value = out_ind_bat[names_ind_batters[i]]


sheet["A29"] = "BATTERS"
sheet["E29"] = "RUNS"
sheet["F29"] = "BALLS"
sheet["G29"] = " 4s "
sheet["H29"] = " 6s "
sheet["I29"] = "  SR  "


sheet["A40"] = "BOWLER"
sheet["C40"] = "OVER"
sheet["D40"] = "MAIDEN"
sheet["E40"] = "RUNS"
sheet["F40"] = "WICKET"
sheet["G40"] = "NO-BALL"
sheet["H40"] = "WIDE"
sheet["I40"] = "ECONOMY"

names_ind_bowlers = []
for count in ind_bowlers.keys():
    names_ind_bowlers.append(count)

for i in range(len(ind_bowlers)):

    sheet.cell(20+i, 1).value = names_ind_bowlers[i]
    sheet.cell(20+i, 3).value = ind_bowlers[names_ind_bowlers[i]][0]
    sheet.cell(20+i, 4).value = ind_bowlers[names_ind_bowlers[i]][1]
    sheet.cell(20+i, 5).value = ind_bowlers[names_ind_bowlers[i]][2]
    sheet.cell(20+i, 6).value = ind_bowlers[names_ind_bowlers[i]][3]
    sheet.cell(20+i, 7).value = ind_bowlers[names_ind_bowlers[i]][4]
    sheet.cell(20+i, 8).value = ind_bowlers[names_ind_bowlers[i]][5]
    sheet.cell(20+i, 9).value = ind_bowlers[names_ind_bowlers[i]][6]
    ind_bowlers_total += ind_bowlers[names_ind_bowlers[i]][2]
    pakistan_fall_of_wickets += ind_bowlers[names_ind_bowlers[i]][3]

ind_total_score = ind_bowlers_total+pak_byes
pak_total_score = pak_bowlers_total+ind_byes



sheet["E27"] = " "+str(ind_total_score) + " - " + str(indian_fall_of_wickets)

sheet["F27"] = str(over_ind)

Eone = " "+str(pak_total_score) + " - " + str(pakistan_fall_of_wickets)

Fone = str(over_pak)

wb.save("Scoreboard.xlsx")

df = pd.read_excel('Scoreboard.xlsx')

df = df.set_axis(['PAKISTAN', ' INNINGS'] + [" ", " ",
                 Eone, Fone, " ", " ", " "], axis='columns')
try:
    df.to_csv('Scorecard.csv', index=False)
except:
    print("output file not saved")


#################         final output is saved in file scorecard.csv      ##################
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
