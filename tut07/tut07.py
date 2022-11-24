from datetime import datetime
start_time = datetime.now()
import math
import glob
import os
import openpyxl
from pathlib import Path
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill
from openpyxl import workbook
import pandas as pd


file_name=glob.glob("input\\"+"*.xlsx")

for i in file_name:
	variable1=Path(i).stem
	file = pd.read_excel(i)
	MOD = 5000
   # finding the average value of U, V AND W
	average_u = file['U'].mean()
	average_v = file['V'].mean()
	average_w = file['W'].mean()

    # subtarcting the value from average values and adding to csv file
	file.at[0, "U Avg"] = round(average_u,3)
	file.at[0, "W Avg"] = round(average_w,3)
	file.at[0, "V Avg"] = round(average_v,3)



	file["U'=U - U avg"] = round((file['U']-average_u),3)
	file["W'=W - W avg"] = round((file['W']-average_w),3)
	file["V'=V - V avg"] = round((file['V']-average_v),3)


    # +ve, +ve 1st quadrant
	# -ve , +ve 2nd quadrant
	# -ve, -ve 3rd quadrant
	# +ve, -ve 4th quadrant

	octantList = []

 
	try:
		for x in range(len(file)):

			if (file["V'=V - V avg"][x] > 0 and file["U'=U - U avg"][x] > 0):
				if (file["W'=W - W avg"][x] >= 0):
					octantList.append(1)
				else:
					octantList.append(-1)

			elif (file["V'=V - V avg"][x] > 0 and file["U'=U - U avg"][x] < 0):
				if (file["W'=W - W avg"][x] >= 0):
					octantList.append(2)
				else:
					octantList.append(-2)

			elif (file["V'=V - V avg"][x] < 0 and file["U'=U - U avg"][x] < 0):
				if (file["W'=W - W avg"][x] >= 0):
					octantList.append(3)
				else:
					octantList.append(-3)

			else:
				if (file["W'=W - W avg"][x] >= 0):
					octantList.append(4)
				else:
					octantList.append(-4)
	except:
		print("Error! could not read excel file")

	file["Octant"] = octantList      
	file[''] = ''
    
	file.at[2,' '] = 'Mod ' + str(MOD)
	file.at[1, "Overall Octant Count"] = "Octant ID"  
	file.at[2, "Overall Octant Count"] = "Overall count"


	file.at[1,'  '] ='1'
	file.at[1,'   ']='-1'
	file.at[1,'    ']='2'
	file.at[1,'     ']='-2'
	file.at[1,'      ']='3'
	file.at[1,'       ']='-3'
	file.at[1,'        ']='4'
	file.at[1,'         ']='-4'

	   
	total_range = math.ceil(len(file)/MOD)            
	start = 0


	for i in range(total_range):


		if (i == total_range-1):
			file.at[i+3, "Overall Octant Count"] = "{}-{}".format(start, len(file)-1)

			file.at[i+3,'  '] = file["Octant"].iloc[start:len(file)].value_counts()[1]
			file.at[i+3,'   '] = file["Octant"].iloc[start:len(file)].value_counts()[-1]
			file.at[i+3,'    '] = file["Octant"].iloc[start:len(file)].value_counts()[2]
			file.at[i+3,'     '] = file["Octant"].iloc[start:len(file)].value_counts()[-2]
			file.at[i+3,'      '] = file["Octant"].iloc[start:len(file)].value_counts()[3]
			file.at[i+3,'       '] = file["Octant"].iloc[start:len(file)].value_counts()[-3]
			file.at[i+3,'        '] = file["Octant"].iloc[start:len(file)].value_counts()[4]
			file.at[i+3,'         '] = file["Octant"].iloc[start:len(file)].value_counts()[-4]
			continue

		file.at[i+3, "Overall Octant Count"] = "{}-{}".format(start, start+MOD-1)

		file.at[i+3, '  '] = file["Octant"].iloc[start:start+MOD].value_counts()[1]
		file.at[i+3, '   '] = file["Octant"].iloc[start:start + MOD].value_counts()[-1]
		file.at[i+3, '    '] = file["Octant"].iloc[start:start+MOD].value_counts()[2]
		file.at[i+3, '     '] = file["Octant"].iloc[start:start + MOD].value_counts()[-2]
		file.at[i+3, '      '] = file["Octant"].iloc[start:start+MOD].value_counts()[3]
		file.at[i+3, '       '] = file["Octant"].iloc[start:start + MOD].value_counts()[-3]
		file.at[i+3, '        '] = file["Octant"].iloc[start:start+MOD].value_counts()[4]
		file.at[i+3, '         '] = file["Octant"].iloc[start:start + MOD].value_counts()[-4]
		start = start+MOD
	


	file.at[2, '  '] = file["Octant"].value_counts()[1]
	file.at[2, '   '] = file["Octant"].value_counts()[-1]
	file.at[2, '    '] = file["Octant"].value_counts()[2]
	file.at[2, '     '] = file["Octant"].value_counts()[-2]
	file.at[2, '      '] = file["Octant"].value_counts()[3]
	file.at[2, '       '] = file["Octant"].value_counts()[-3]
	file.at[2, '        '] = file["Octant"].value_counts()[4]
	file.at[2, '         '] = file["Octant"].value_counts()[-4]
	
	
	
	file.at[1, '          '] = "Rank of 1"
	file.at[1, '           '] = "Rank of -1"
	file.at[1, '            '] = "Rank of 2"
	file.at[1, '             '] = "Rank of -2"
	file.at[1, '              '] = "Rank of 3"
	file.at[1, '               '] = "Rank of -3"
	file.at[1, '                '] = "Rank of 4"
	file.at[1, '                 '] = "Rank of -4"

	file.at[1, '                  '] = "Rank1 Octant ID"
	file.at[1, '                   '] = 'Rank 1 Octant Name'
	

	count_list =[]

	for i in range(8):
		count_list.append(file.at[2, ' '*(i+2)])


	dict_count = {count_list[0]: 0, count_list[1]: 0, count_list[2]: 0, count_list[3]: 0, count_list[4]: 0, count_list[5]: 0, count_list[6]: 0, count_list[7]: 0}
	start_1 =8

    # print(max(count_list))
	#assigning values from 1 to 8
	for i in range(8):
		mini = min(count_list)
		dict_count[mini] = start_1
		count_list.remove(mini)
		start_1-=1
	# print(dict_count)
	# print the values in xlsx in file
	start_cout =1
	for values in dict_count.values():
		# print(values)
		file.at[2, ' '*(start_cout+9)] = values
		start_cout+=1
	for i in range(8):
		if(file.at[2, ' '*(i+10)] ==1):
				if(i==0):
					file.at[2, '                  '] = 1
					file.at[2, '                   '] = "Internal outward interaction"
				if(i==1):
					file.at[2, '                  '] = -1
					file.at[2, '                   '] = "External outward interaction"
				if(i==2):
					file.at[2, '                  '] = 2
					file.at[2, '                   '] = "External Ejection"
				if(i==3):
					file.at[2, '                  '] = -2
					file.at[2, '                   '] = "Internal Ejection"
				if(i==4):
					file.at[2, '                  '] = 3
					file.at[2, '                   '] = "External inward interaction"
				if(i==5):
					file.at[2, '                  '] = -3
					file.at[2, '                   '] = "Internal inward interaction"
				if(i==6):
					file.at[2, '                  '] = 4
					file.at[2, '                   '] = "Internal sweep"
				if(i==7):
					file.at[2, '                  '] = -4
					file.at[2, '                   '] = "External sweep"

	
# for mod
	def priority_order(mod =5000):
		total_range = math.ceil(len(file)/mod)
		count_list1 =[]
		start_cout_1 =0
		for i in range(total_range):
			for i in range(8):
				count_list1.append(file.at[start_cout_1+3, ' '*(i+2)])
		
			List_count_dict_1 = {count_list1[0]: 0, count_list1[1]: 0, count_list1[2]: 0, count_list1[3]: 0, count_list1[4]: 0, count_list1[5]: 0, count_list1[6]: 0, count_list1[7]: 0}

			start1 =8
			for i in range(8):
				mini1 = min(count_list1)
				List_count_dict_1[mini1] = start1
				count_list1.remove(mini1)
				start1-=1

			start_cout =1
			for values in List_count_dict_1.values():      
				file.at[start_cout_1+3,' '*(start_cout+9)] = values
				start_cout+=1

		    ## rank1 octant ID
			for i in range(8):
				if(file.at[start_cout_1+3,' '*(i+10)] ==1):
					if(i==0):
						file.at[start_cout_1+3, '                  '] = 1
						file.at[start_cout_1+3, '                   '] = "Internal outward interaction"
					if(i==1):
						file.at[start_cout_1+3, '                  '] = -1
						file.at[start_cout_1+3, '                   '] = "External outward interaction"
					if(i==2):
						file.at[start_cout_1+3, '                  '] = 2
						file.at[start_cout_1+3, '                   '] = "External Ejection"
					if(i==3):
						file.at[start_cout_1+3, '                  '] = -2
						file.at[start_cout_1+3, '                   '] = "Internal Ejection"
					if(i==4):
						file.at[start_cout_1+3, '                  '] = 3
						file.at[start_cout_1+3, '                   '] = "External inward interaction"
					if(i==5):
						file.at[start_cout_1+3, '                  '] = -3
						file.at[start_cout_1+3, '                   '] = "Internal inward interaction"
					if(i==6):
						file.at[start_cout_1+3, '                  '] = 4
						file.at[start_cout_1+3, '                   '] = "Internal sweep"
					if(i==7):
						file.at[start_cout_1+3, '                  '] = -4
						file.at[start_cout_1+3, '                   '] = "External sweep"
			start_cout_1 +=1


	priority_order(MOD)
	total_range = math.ceil(len(file)/MOD)
	numb_list =["Octant ID", 1,  -1, 2, -2, 3, -3, 4, -4]
	octant_name_list =["Octant Name","Internal outward interaction", "External outward interaction", "External Ejection", "internal Ejection", "External inward interaction", "Internal inward interaction", "Internal sweep", "External sweep"]
	count_of_no = []

	for i in range(total_range):
		count_of_no.append(file.at[2+i,'                  '])
	for i in range(9):
		file.at[5+total_range+i+1, '                '] = numb_list[i]
		file.at[5+total_range+i+1, '                 '] = octant_name_list[i]
		
	file.at[6+total_range, '                  '] = "Count of Rank 1 Mod Values"
	no_list = [1, -1, 2, -2, 3, -3, 4, -4]

	for i in range(8):
		file.at[7+total_range+i, '                  '] = count_of_no.count(no_list[i])
	


	file['                    '] = ''
	file.at[2,'                     '] = 'From'
		
	total_range = math.ceil(len(file)/MOD)
	total_range = total_range+5

	file.at[1, "Overall transition count"] = 'Octant #'

	row_index = [1, -1, 2, -2, 3, -3, 4, -4]
	index = [1, -1, 2, -2, 3, -3, 4, -4]
	file.at[0, '                      '] = "To"
	for i in range(1, len(index)+1):
		file.at[i+1, "Overall transition count"] = index[i-1]

	file.at[1, '                      '] = "1"
	file.at[1, '                       '] = "-1"
	file.at[1, '                        '] = "2"
	file.at[1, '                         '] = "-2"
	file.at[1, '                          '] = "3"
	file.at[1, '                           '] = "-3"
	file.at[1, '                            '] = "4"
	file.at[1, '                             '] = "-4"

	list_total =[
							[0, 0, 0, 0, 0, 0, 0, 0],
							[0, 0, 0, 0, 0, 0, 0, 0],
							[0, 0, 0, 0, 0, 0, 0, 0],
							[0, 0, 0, 0, 0, 0, 0, 0],
							[0, 0, 0, 0, 0, 0, 0, 0],
							[0, 0, 0, 0, 0, 0, 0, 0],
							[0, 0, 0, 0, 0, 0, 0, 0],
							[0, 0, 0, 0, 0, 0, 0, 0]
							]


	for i in range(len(file)-1):
		value1 = file["Octant"][i]
		value2 = file["Octant"][i+1]
		value1_index = index.index(value1)
		# print(value1_index)
		value2_index = row_index.index(value2)
		list_total[value1_index][value2_index] += 1
	# mid transition count
	#row_index = [1, -1, 2, -2, 3, -3, 4, -4]
	#index = ['Count', 1, -1, 2, -2, 3, -3, 4, -4]
	# print(s)


	for i in range(8):
		file.at[2+i, '                      '] = list_total[i][0]
		file.at[2+i, '                       '] = list_total[i][1]
		file.at[2+i, '                        '] = list_total[i][2]
		file.at[2+i, '                         '] = list_total[i][3]
		file.at[2+i, '                          '] = list_total[i][4]
		file.at[2+i, '                           '] = list_total[i][5]
		file.at[2+i, '                            '] = list_total[i][6]
		file.at[2+i, '                             '] = list_total[i][7]
		
	total_range = math.ceil(len(file)/MOD)
	var = 13
	initial = 0
	while(total_range>0):
		file.at[var, "Overall transition count"] = "Mod Transition Count"

		final = initial + MOD
		if(final>len(file)):
			final = len(file)-2 
			
		file.at[var+1, "Overall transition count"] = "{}-{}".format(initial, final-1)

		if(initial!=0):
			file.at[var+1, "Overall transition count"] = "{}-{}".format(initial, final-1)
		if(total_range==1):
			file.at[var+1, "Overall transition count"] ="{}-{}".format(initial, len(file)-1)


		file.at[var+1, '                      '] = "To"
		file.at[var+3, '                     '] = "From"
		file.at[var+2,"Overall transition count"] = 'octant##'

		for i in range(1,len(index)+1):
			file.at[var+2+i, "Overall transition count"] = index[i-1]

		file.at[var+2, '                      '] = "1"
		file.at[var+2, '                       '] = "-1"
		file.at[var+2, '                        '] = "2"
		file.at[var+2, '                         '] = "-2"
		file.at[var+2, '                          '] = "3"
		file.at[var+2, '                           '] = "-3"
		file.at[var+2, '                            '] = "4"
		file.at[var+2, '                             '] = "-4"


		list_total1=[
						[0, 0, 0, 0, 0, 0, 0, 0],
						[0, 0, 0, 0, 0, 0, 0, 0],
						[0, 0, 0, 0, 0, 0, 0, 0],
						[0, 0, 0, 0, 0, 0, 0, 0],
						[0, 0, 0, 0, 0, 0, 0, 0],
						[0, 0, 0, 0, 0, 0, 0, 0],
						[0, 0, 0, 0, 0, 0, 0, 0],
						[0, 0, 0, 0, 0, 0, 0, 0]
					]

		for i in range(initial,final):
			value1 = file["Octant"][i]
			value2 = file["Octant"][i+1]
			value1_index = index.index(value1)
			value2_index = row_index.index(value2)
			list_total1[value1_index][value2_index] += 1
		
		
		t = var+3
		for i in range(8):
			file.at[t, '                      '] = list_total1[i][0]
			file.at[t, '                       '] = list_total1[i][1]
			file.at[t, '                        '] = list_total1[i][2]
			file.at[t, '                         '] = list_total1[i][3]
			file.at[t, '                          '] = list_total1[i][4]
			file.at[t, '                           '] = list_total1[i][5]
			file.at[t, '                            '] = list_total1[i][6]
			file.at[t, '                             '] = list_total1[i][7]
			t += 1
		initial = final
		var = var+13
		total_range = total_range-1



	
	file['                              '] = ''
	file["Octant"] = octantList    
	variablelist =["1","-1","2","-2","3","-3","4","-4"]
	removestringlist =[1, -1, 2, -2, 3, -3, 4, -4]
	initialcount =0
	valueCount = { -4: 0,-3: 0,-2: 0,-1: 0, 1: 0, 2: 0, 3: 0, 4: 0  }    

	
	for i in range(-4, 5):
		if (i == 0):
			continue
		flag = 0                                          
		for j in range(len(file)-1):
			if (file.at[j, "Octant"] == file.at[j+1, "Octant"] and file.at[j, "Octant"] == i):
				flag += 1
			else:
				valueCount[i] = max(valueCount[i], flag+1)
				flag = 0

	
	list_1=[]
	list_to_1=[]

	list_mod1=[]
	list_to_mod1 =[]

	list_2=[]
	list_to_2 =[]

	list_mod2=[]
	list_to_mod2 =[]

	list_3=[]
	list_to_3=[]

	list_mod3=[]
	list_to_mod3=[]

	list_4=[]
	list_to_4= []

	list_mod4=[]
	list_to_mod4 =[]

	countLength = {-4: 0, -3: 0,-2: 0, -1: 0,1: 0,2: 0,3: 0, 4: 0  }

	
	for i in range(-4, 5):
		if (i == 0):
			continue
		flag = 0
		for j in range(len(file)-1):
			if (file.at[j, "Octant"] == file.at[j+1, "Octant"] and file.at[j, "Octant"] == i):
				flag += 1
				

			else:
				if (flag+1 == valueCount[i]):
					if(i==1):
						list_1.append(file["T"][j-flag])
						list_to_1.append(file["T"][j])
					if(i==-1):
						list_mod1.append(file["T"][j-flag])
						list_to_mod1.append(file["T"][j])
					if(i==2):
						list_2.append(file["T"][j-flag])
						list_to_2.append(file["T"][j])
					if(i==-2):
						list_mod2.append(file["T"][j-flag])
						list_to_mod2.append(file["T"][j])
					if(i==3):
						list_3.append(file["T"][j-flag])
						list_to_3.append(file["T"][j])
					if(i==-3):
						list_mod3.append(file["T"][j-flag])
						list_to_mod3.append(file["T"][j])
					if(i==4):
						list_4.append(file["T"][j-flag])
						list_to_4.append(file["T"][j])
					if(i==-4):
						list_mod4.append(file["T"][j-flag])
						list_to_mod4.append(file["T"][j])

					countLength[i] += 1
				flag= 0

	
	file.at[1,'Longest Subsquence Length'] ="Octant ##" 


	file.at[2,'Longest Subsquence Length'] = "1"
	file.at[3,'Longest Subsquence Length'] = "-1"
	file.at[4,'Longest Subsquence Length'] = "2"
	file.at[5,'Longest Subsquence Length'] = "-2"
	file.at[6,'Longest Subsquence Length'] = "3"
	file.at[7,'Longest Subsquence Length'] = "-3"
	file.at[8,'Longest Subsquence Length'] = "4"
	file.at[9,'Longest Subsquence Length'] = "-4"


	
	countLength_count = 0
	for i in range(8):
		file.at[i+2, '                               ']= valueCount[removestringlist[i]]
		file.at[i+2, '                                '] = countLength[removestringlist[i]]
		countLength_count += countLength[removestringlist[i]]


	file['                                 '] = ''


	
	file["Longest Subsquence Length with Range"] = " "
	file['                                  '] = " "
	file['                                   '] =" "

	variablelist =["1","-1","2","-2","3","-3","4","-4"]
	removestringlist =[1, -1, 2, -2, 3, -3, 4, -4]
	initialcount =0
	file.at[1,"Longest Subsquence Length with Range"]="Octant ###"
	file.at[1,'                                  '] = "Longest Subsquence Length"
	file.at[1,'                                   '] ="Count"

	for i in range(8):
		file.at[initialcount+2,"Longest Subsquence Length with Range"] = variablelist[i]
		file.at[initialcount+2, '                                  '] = valueCount[removestringlist[i]]
		file.at[initialcount+2, '                                   '] = countLength[removestringlist[i]]

		file.at[3+initialcount, "Longest Subsquence Length with Range"] ="Time"
		file.at[3+initialcount, '                                  '] = "From"
		file.at[3+initialcount, '                                   '] = "To"

		for j in range(countLength[removestringlist[i]]):
			if(removestringlist[i]==1):
				file.at[4+initialcount+j,'                                  ']=list_1[j]
				file.at[4+initialcount+j,'                                   ']=list_to_1[j]
			elif(removestringlist[i]==2):
				file.at[4+initialcount+j,'                                  ']=list_2[j]
				file.at[4+initialcount+j,'                                   ']=list_to_2[j]
			elif(removestringlist[i]==3):
				file.at[4+initialcount+j,'                                  ']=list_3[j]
				file.at[4+initialcount+j,'                                   ']=list_to_3[j]
			elif(removestringlist[i]==4):
				file.at[4+initialcount+j,'                                  ']=list_4[j]
				file.at[4+initialcount+j,'                                   ']=list_to_4[j]
			elif(removestringlist[i]==-1):
				file.at[4+initialcount+j,'                                  ']=list_mod1[j]
				file.at[4+initialcount+j,'                                   ']=list_to_mod1[j]
			elif(removestringlist[i]==-2):
				file.at[4+initialcount+j,'                                  ']=list_mod2[j]
				file.at[4+initialcount+j,'                                   ']=list_to_mod2[j]
			elif(removestringlist[i]==-3):
				file.at[4+initialcount+j,'                                  ']=list_mod3[j]
				file.at[4+initialcount+j,'                                   ']=list_to_mod3[j]
			elif(removestringlist[i]==-4):
				file.at[4+initialcount+j,'                                  ']=list_mod4[j]
				file.at[4+initialcount+j,'                                   ']=list_to_mod4[j]   
				
		initialcount += countLength[removestringlist[i]]+2
	FILE_NAME1 = os.path.join("output/" +variable1+'_vel_octant_analysis_mod_'+str(MOD)+'.xlsx')
	file.to_excel(FILE_NAME1,index = False)
	

	wb = openpyxl.load_workbook(FILE_NAME1)
	sheet = wb['Sheet1']
    

	###############################
	# code for adding color
	top = Side(border_style = 'thin', color = '000000')
	bottom = Side(border_style = 'thin', color = '000000')  
	left = Side(border_style = 'thin', color = '000000')
	right = Side(border_style = 'thin', color = '000000')
	border = Border(top = top, bottom = bottom, left = left, right = right)
	colour = PatternFill( start_color ='FFFF00',end_color ='FFFF00', fill_type='solid' )
	total_range = math.ceil(len(file)/MOD)
	
	for Row in range(3, total_range+5):
		for Column in range(14,33):
			sheet.cell(row=Row, column = Column).border = border
	
	for Row in range(4,total_range+5):
		for Column in range(23,31):   
			if sheet.cell(row =Row,column=Column).value == 1:              
				sheet.cell(row = Row ,column = Column).fill = colour

	for Row in range(total_range+8, total_range+17):
		for Column in range(29,32):
			sheet.cell(row = Row ,column = Column).border = border
	for Row in range(3, total_range+8):
		for Column in range(35,44):
			sheet.cell(row = Row ,column = Column).border = border
	a =[] 
	for Row in range(4, 12):
		for Column in range(36,44):
			a.append(sheet.cell(row = Row ,column = Column).value)
		z = max(a)                                              
		for Column in range(36,44):
			if sheet.cell(row = Row ,column = Column).value == z:
				sheet.cell(row = Row ,column = Column).fill = colour
		a = []  

		
		
	for i in range(total_range):
		for Row in range(17+i*13, 26+i*13):
			for Column in range(35,44):
				sheet.cell(row = Row ,column = Column).border = border
	b = []
	for i in range(total_range):
		for Row in range(18+i*13, 26+i*13):
			for Column in range(36,44):
				b.append(sheet.cell(row = Row ,column = Column).value)     
			z = max(b)
			for Column in range(36,44):
				if sheet.cell(row = Row ,column = Column).value == z:
					sheet.cell(row = Row ,column = Column).fill = colour
			b = []   



	for Row in range(3, 12):
		for Column in range(45,48):
			sheet.cell(row = Row ,column = Column).border = border
	for Row in range(3, 20+countLength_count):
		for Column in range(49,52):
			sheet.cell(row = Row ,column = Column).border = border

	wb.save(FILE_NAME1)
	
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))